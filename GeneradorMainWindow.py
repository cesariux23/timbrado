import sys, os, re, shutil
from heapq import merge
from openpyxl import *
from datetime import date
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QLabel, QFileDialog
from PyQt5 import uic
from idQuincena import *
from datetime import datetime, date, time, timedelta

Ui_generadorMainWindow, QtBaseClass = uic.loadUiType("ui/generador.ui")

class GeneradorMainWindow(QMainWindow):

    REGISTRO_PATRONAL = '06030087'

    def __init__(self):
        super(GeneradorMainWindow, self).__init__()
        self.ui = Ui_generadorMainWindow()
        self.ui.setupUi(self)
        self.quincena_actual = IdQuincena(date.today())
        self.quincena_actual.set_quincena(self.quincena_actual.quincena-1)
        self.ui.txtanio.setText(str(self.quincena_actual.anio))
        self.ui.txtqna.setText(str(self.quincena_actual.quincena))
        self.set_quincena()
        self.ui.btnseleccionar.clicked.connect(self.seleccionarArchivo)
        self.ui.btngenerar.clicked.connect(self.generarTimbrado)
        self.ui.txtarchivo.textChanged.connect(self.cambiaArchivo)
        self.ui.txtanio.textChanged.connect(self.cambia_datos_qna)
        self.ui.txtqna.textChanged.connect(self.cambia_datos_qna)
        self.ui.cbtiponomina.currentIndexChanged.connect(self.cambia_tipo)
        self.archivo_catalogo = 'catalogo.xlsx'
        ##Define los campos requeridos para el timbrado
        self.datos = ['RFC', 'CURP', 'CODIGO', 'NSS',
                      'NCUENTA', 'IDPAGO', 'NOEMPEADO', 'ADSCRIPCION', 'NOMBRE',
                      'NOMBRE_PUESTO', 'CORREO', 'TPERCEP', 'TDEDUC', 'TNETO',
                      'FECHAING', 'BASE_SQC', 'BASECONF', 'IDPAGO', 'NCUENTA']
        self.percepciones = ["SUELDO", "QUINQUENIO", "PREVISION SOCIAL", "DESPENSA", "AEICS",
                             "APOYO_CYD", "PRODUCTIVIDAD", "TRES", "COMPENSACION", "GUARDERIA",
                             "PUNT_MENS", "PUNT_PERF", "PERMISOS", "PUNT_ANUAL",
                             "DIAS", "MENSUAL", "TRIMESTRAL", "TRIMESTRAL2", "PRIMA",
                             "DEVOLUCION", 'DIA_EMP', 'PROD_ANUAL', 'ACREDITAMIENTO',
                             'AGUINALDO', 'ISR_AGUINALDO', 'SEG_RETIRO']
        self.deducciones = ["CUO_SIN", "ISSSTE", "SEG_MEDICO", "SEG_RETIRO",
                            "POTEN", "AHORRO", "ISRHOMO", 'ISR', 'PENSION', 'DAÑOS',
                            'CINCO', "PCP", "PHIP", "AHISA", "OTRO", "CUATRO"]

        self.carpeta = os.path.expanduser('~') + "/documents"
        #se carga el catalogo de percepciones/deducciones
        self._carga_catalogo_()
        self.tipo_nomina = 0
        self.regimen_empleado = '02'
        self.registrop = self.REGISTRO_PATRONAL
        self.plantilla = True

    def cambia_tipo(self, i):
        self.tipo_nomina = i
        if str(i) in '013':
            self.regimen_empleado = '02'
            self.registrop = self.REGISTRO_PATRONAL
            self.plantilla = True
        else:
            #honorarios
            self.regimen_empleado = '09'
            self.registrop = ''
            self.plantilla = False

    def cambia_datos_qna(self):
        try:
            if int(self.ui.txtanio.text()) >= 2015 <= 2020 and int(self.ui.txtqna.text()) >= 1 <= 24:
                self.quincena_actual.anio = int(self.ui.txtanio.text())
                self.quincena_actual.set_quincena(int(self.ui.txtqna.text()))
                self.set_quincena()
                self.fecha_pago = []
                self.fecha_pago.append(self.quincena_actual.fecha_fin.strftime("%d/%m/%Y"))
                self.fecha_pago.append(self.quincena_actual.fecha_inicio.strftime("%d/%m/%Y"))
                self.fecha_pago.append(self.quincena_actual.fecha_fin.strftime("%d/%m/%Y"))
        except ValueError:
            print('Valores no validos')

    #establece la quincena actual
    def set_quincena(self):
        self.ui.txt_descripcion.setText(self.quincena_actual.nombre.upper())

    def seleccionarArchivo(self):
        self.ui.btngenerar.setEnabled(False)
        self.archivo_nomina, _ = QFileDialog.getOpenFileName(self, "Abrir archivo de nomina ( Excel )", self.carpeta, "Archivos de Excel (*.xlsx)")
        self.ui.txtarchivo.setText(str(self.archivo_nomina))
        if(len(self.archivo_nomina)>0):
            self.carpeta = os.path.dirname(self.archivo_nomina);
            self.ui.btngenerar.setEnabled(True)
            #determina el nombre de archivo para calcular la quincena a partir del nombre del archivo
            palabras = self.archivo_nomina.split("/")
            nombre = palabras[len(palabras)-1][:-5]
            if len(nombre) >= 6:
                self.ui.txtanio.setText(nombre[:4])
                self.ui.txtqna.setText(nombre[4:6])
                self.cambia_datos_qna()

    def cambiaArchivo(self):
        print(self.ui.txtarchivo.text())

    def generarTimbrado(self):
        self.ui.txtsalida.setText('Procesando el archivo '+self.archivo_nomina)
        #define la base del folio
        base_folio = self.quincena_actual.id.replace('20', '', 1)+str(self.tipo_nomina)+self.ui.txtenvio.text()
        folio = 1

        #descripcion del pago
        pago = self.ui.txt_descripcion.text()
        #regex para buscar digitos
        p = re.compile('\d+')
        #abre libro para datos generales
        wb_datos = load_workbook("datos_generales.xlsx")
        ws_datos = wb_datos.active


        #abre libro para conceptos
        wb_conceptos = load_workbook("conceptos.xlsx")
        ws_conceptos = wb_conceptos.active

        #abre el libro de excel de la nomina
        wb = load_workbook(filename=self.archivo_nomina, read_only=True, data_only=True)
        #recupera el nombre de las hojas
        hojas = wb.get_sheet_names()
        for nombre_hoja in wb.get_sheet_names():
            hoja = wb[nombre_hoja]
            print(nombre_hoja)
            #limpia arreglos
            indicedatos = {}
            indiceconceptos = {}
            self.ui.txtsalida.append('--Procesando hoja ' + nombre_hoja + ': ' + str(hoja.max_row) + ' filas.')
            ##recorre los encabezados
            for c in range(1, hoja.max_column+1):
                try:
                    celda = hoja.cell(row=1, column=c)
                    if not celda.value is None:
                        #limpia los valores
                        valor = self.limpiaValor(celda.value).upper()
                        #determina los indices
                        encontrado = False
                        #datos generaleS
                        for dato in self.datos:
                            if dato == valor:
                                indicedatos[dato] = c
                                encontrado = True
                        if not encontrado:
                            for perc in self.catalogo:
                                if perc == valor:
                                    indiceconceptos[perc] = c
                                    encontrado = True
                        if not encontrado:
                            print('Se omite el campo '+valor)
                            self.ui.txtsalida.append('\t** Se omite el campo '+valor)
                except Exception as e:
                    print(e)
                    print("No existe la celda")
            ##se notifica si no existe el campo RFC y curp
            if 'RFC' not in indicedatos:
                self.ui.txtsalida.append('\t***No existe la columna RFC')
            #determina si es plantilla o honorarios
            es_plantilla = False
            tipo_contrato = '09'
            if 'CODIGO' in indicedatos and self.plantilla:
                es_plantilla = True
                tipo_contrato = '01'
            print(es_plantilla)
            #recorre filas para los empleados
            for row in hoja.iter_rows(row_offset=1):
                rfc = row[indicedatos['RFC']-1].value
                if not rfc is None:
                    folio_empleado = base_folio+str(folio).zfill(4)
                    folio += 1
                    #acumulado del excento
                    exento = 0.0
                    #consecutivo de mov
                    mov = 1
                    #recorre los conceptos
                    for pr in indiceconceptos:
                        valor_concepto = row[indiceconceptos[pr]-1].value
                        if not valor_concepto is None:
                            valor_real = round(float(str(valor_concepto).replace(',', '')), 2)
                            if valor_real > 0.0:
                                descripcion = self.catalogo[pr]
                                #Se agrega al concentrado de conceptos
                                concepto = [folio_empleado, str(mov)] + list(descripcion[:4])
                                mov += 1
                                #se determina si la percepcion graba o no
                                try:
                                    valor_excento = descripcion[4]
                                    if valor_excento:
                                        concepto.append(0.0)
                                        concepto.append(valor_real)
                                        #suma las percepciones
                                        if descripcion[0]=='1':
                                            exento += valor_real
                                    else:
                                        concepto.append(valor_real)
                                        concepto.append(0.0)
                                except:
                                    concepto.append(0.0)
                                    concepto.append(valor_real)
                                #se agrega el registro a la hoja de excel
                                ws_conceptos.append(concepto)
                                #se agrega Aportaciones Federales
                                #ws_conceptos.append((folio_empleado,str(mov),"1","P41", "INGRESOS FEDERALES", "041",0,0))

                    #se general el anexo de datos generales
                    datos_empleado = [folio_empleado, "IVE"]
                    #determina si es plantilla o honorarios
                    # es_plantilla = False
                    # tipo_contrato = '09'
                    # if 'CODIGO' in indicedatos and self.plantilla:
                    #     es_plantilla = True
                    #     tipo_contrato = '01'
                    #datos estaticos
                    percepciones = round(row[indicedatos['TPERCEP']-1].value, 2)
                    deducciones = round(row[indicedatos['TDEDUC']-1].value, 2)
                    isr = round(row[indiceconceptos['ISR']-1].value, 2)
                    descuentos = deducciones - isr
                    neto = round(percepciones-deducciones,2)
                    nombre = row[indicedatos['NOMBRE']-1].value
                    curp = row[indicedatos['CURP']-1].value
                    nss = row[indicedatos['NSS']-1].value
                    fecha_ingreso = row[indicedatos['FECHAING']-1].value
                    baseconf = row[indicedatos['BASECONF']-1].value
                    numero_emp = row[indicedatos['NOEMPEADO']-1].value
                    if not numero_emp:
                        numero_emp="0000"
                    adscripcion = self.limpiaValor(row[indicedatos['ADSCRIPCION']-1].value)
                    puesto = self.limpiaValor(row[indicedatos['NOMBRE_PUESTO']-1].value)
                    tipo_pago = row[indicedatos['IDPAGO']-1].value
                    cuenta_deposito = row[indicedatos['NCUENTA']-1].value
                    sueldo_aport = round(row[indicedatos['BASE_SQC']-1].value, 2)
                    correo = row[indicedatos['CORREO']-1].value
                    if correo is None:
                        correo = 'ver_rechum@inea.gob.mx'

                    #subtotal, descuentos, total
                    datos_empleado += [percepciones, deducciones, neto]
                    datos_empleado += ['91030', rfc, nombre]
                    #valor unitario, importe
                    datos_empleado += [percepciones, percepciones]
                    #tipo de nomina
                    datos_empleado += ['O']
                    #fecha pago, inicio, fin
                    datos_empleado += self.fecha_pago
                    #dias pagados
                    datos_empleado.append(15.0)
                    #percepciones, deducciones, otros pagos
                    datos_empleado += [percepciones, deducciones, 0]
                    #nodo emisor
                    #registro patronal , origen de recurso, recurso propio
                    datos_empleado += [self.registrop, 'IF', 0.0]
                    #nodo receptor
                    datos_empleado.append(curp)
                    #NSS
                    datos_empleado.append(nss)
                    #Fecha inicio relacion laboral
                    datos_empleado.append(fecha_ingreso.strftime("%d/%m/%Y"))
                    #antiguedad
                    ant = self.calcular_antiguedad(fecha_ingreso)
                    datos_empleado.append(ant)
                    #tipo contrato
                    datos_empleado.append(tipo_contrato)
                    #sindicalizado
                    sindicalizado = self.es_sindicalizado(baseconf)
                    datos_empleado.append(sindicalizado)
                    #jornada
                    datos_empleado.append('01')
                    #regimen
                    datos_empleado.append(self.regimen_empleado)
                    #puesto
                    datos_empleado.append(numero_emp)
                    datos_empleado.append(adscripcion)
                    datos_empleado.append(puesto)
                    #riesgo, periodicidad
                    datos_empleado += ['1', '04']
                    #banco, cuenta
                    banco = self.banco(tipo_pago)
                    #cuenta
                    if banco:
                        datos_empleado.append(banco)
                        datos_empleado.append(cuenta_deposito)
                    else:
                        datos_empleado += ['', '']
                    #sueldo base Aportacion
                    datos_empleado.append(sueldo_aport)
                    #salario diario
                    datos_empleado.append(round(percepciones/15, 2))
                    #entidad
                    datos_empleado.append('VER')
                    #sueldos
                    datos_empleado += [percepciones, '', '', percepciones-exento, exento]
                    #otras deducciones, impuestos retenidos
                    datos_empleado += [descuentos, isr]
                    #horas extras
                    datos_empleado += ['', '', '', '', '']
                    #saldos a favor
                    datos_empleado += ['', '', '', '', '', '']
                    #extra
                    datos_empleado.append('')
                    #correo
                    datos_empleado.append(correo)
                    datos_empleado.append('')
                    #descripcion del pago
                    datos_empleado.append(pago)
                    ws_datos.append(datos_empleado)

        #se guarda el libro conceptos
        try:
            os.remove(self.carpeta+"/conceptos.xlsx")
        except:
            pass
        wb_conceptos.save(self.carpeta+"/conceptos.xlsx")

        #se guarda el libro de datos
        try:
            os.remove(self.carpeta+"/datos_generales.xlsx")
        except:
            pass
        wb_datos.save(self.carpeta+"/datos_generales.xlsx")

    def es_sindicalizado(self, x):
        self.es_base = False
        if x == 'B':
            self.es_base = True
            return 'Sí'
        else:
            return ''

    def banco(self, x):
        if x == 1:
            #BANAMEX
            return '002'
        elif x== 2:
            #HSBC
            return '021'
        elif x == 4:
            return '012'
        else:
            return False

    def actualiza_base(self, fila, columna, valor):
        f=self.hojabase[fila]
        celda=f[columna].value=valor
    
    def calcular_antiguedad(self, fecha_ing):
        fecha_q = self.quincena_actual.fecha_fin
        diferencia = datetime.combine(fecha_q, datetime.min.time()) -fecha_ing
        return "P"+str(int(diferencia.days/7))+"W"

    def cierra_base(self):
        self.wbbase.save('base.xlsx')

    def limpiaValor(self,valor):
        return valor.replace(","," ").replace(".","").replace("´","").replace("/","").replace("Í","I").replace("Á","A").replace("É","E").replace("Ó","O").replace("Ú","U").replace("(","").replace(")","").strip()

    def _carga_catalogo_(self):
        wbbase = load_workbook(filename=self.archivo_catalogo)
        hojabase = wbbase.worksheets[0]
        self.catalogo = {}
        count = 0
        for row in hojabase.iter_rows(row_offset=2):
            columna = row[0].value
            if not columna is None:
                self.catalogo[columna] = (
                    str(row[1].value),           #0 tipo
                    str(row[2].value),      #1 clave
                    str(row[3].value),      #2 descripcion
                    str(row[4].value),      #3 tipo_sat
                    int(row[5].value)       #4 exento
                )
                count += 1
        print('\t {0} registros -> Ok.'.format(count))
